import pandas as pd
from openpyxl import load_workbook
import os

# Folder containing the excel files
folder_path = r'C:\\Users\\cleung9\\Desktop\\Relocation\\Wave7'

# List all excel files in the folder
excel_files = [os.path.join(folder_path, file) for file in os.listdir(folder_path) if file.endswith('.xlsx')]

# Create an empty dataframe to store the combined data
combined_df = pd.DataFrame()


# Function to ensure unique column names
def deduplicate_columns(columns):
    seen = set()
    for i, col in enumerate(columns):
        if col in seen:
            columns[i] = f"{col}_{i}"
        seen.add(columns[i])
    return columns


# Iterate over each excel file and append its content to the combined dataframe
for file in excel_files:
    wb = load_workbook(file)
    if 'Port mapping (IP)' in wb.sheetnames:
        sheet = wb['Port mapping (IP)']
        data = sheet.values
        columns = next(data)
        columns = deduplicate_columns(list(columns))  # Ensure unique column names
        df = pd.DataFrame(data, columns=columns)
        df['Source File'] = file  # Add a column for the source file

        combined_df = pd.concat([combined_df, df], ignore_index=True)

# Save the combined dataframe to a new excel file
combined_df.to_excel('combined_file_with_source.xlsx', index=False)

print("The excel files have been successfully combined into 'combined_file_with_source.xlsx'.")