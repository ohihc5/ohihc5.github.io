import pandas as pd
import glob
import os
from openpyxl import load_workbook

folder_path = r"C:\Users\cleung9\Desktop\Relocation\Wave1"
pattern = os.path.join(folder_path, "[!~$]*.xlsx")
excel_files = glob.glob(pattern)
sheet_names = ["Port mapping (IP)", "Port mapping (SAN Switch)"]

combined_data = {sheet: [] for sheet in sheet_names}
for file in excel_files:
    try:
        wb = load_workbook(file)
        for ws in wb.worksheets:
            for col_dim in ws.column_dimensions.values():
                col_dim.hidden = False
            for row_idx in range(1, ws.max_row + 1):
                ws.row_dimensions[row_idx].hidden = False
            ws.auto_filter.ref = None
        wb.save(file)

        xls = pd.ExcelFile(file, engine="openpyxl")
        filename = os.path.basename(file)
        # Project Owner substring: up to first '_', plus "_Wave1"
        project_owner = filename.split("_")[0] + "_Wave1"
        for sheet in sheet_names:
            if sheet in xls.sheet_names:
                df = pd.read_excel(file, sheet_name=sheet, engine="openpyxl")
                df.insert(0, "File Version", filename)
                df.insert(1, "Project Owner", project_owner)
                df.insert(2, "Patch Panel Connection Ref.", "")
                combined_data[sheet].append(df)
    except Exception as e:
        print(f"Skipping {os.path.basename(file)}: {e}")

for sheet, dfs in combined_data.items():
    dfs_valid = [df for df in dfs if not df.empty]
    combined_data[sheet] = pd.concat(dfs_valid, ignore_index=True) if dfs_valid else pd.DataFrame()


output_path = os.path.join(folder_path, "DCNI - Consolidated Patching Record (Wave1).xlsx")
with pd.ExcelWriter(output_path, engine="xlsxwriter") as writer:
    for sheet in sheet_names:
        combined_data[sheet].to_excel(writer, sheet_name=sheet, index=False)

print("Combined workbook written to:", output_path)
