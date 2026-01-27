import re
import sys
import os
from openpyxl import load_workbook
from openpyxl.styles import Font

# ==============================================================================
# SECTION 1: REGEX PATTERN COMPILATION
# ==============================================================================

REGEX_SYSNAME = re.compile(r"^\s*sysname\s+(\S+)")
REGEX_INTERFACE_START = re.compile(r"^\s*interface\s+([\w\/\.\-]+)")
REGEX_DESCRIPTION = re.compile(r"^\s*description\s+(.+)")
REGEX_LINK_TYPE = re.compile(r"^\s*port link-type\s+(access|trunk|hybrid)")
REGEX_DEFAULT_VLAN = re.compile(r"^\s*port default vlan\s+(\d+)")
REGEX_TRUNK_ALLOW_VLAN = re.compile(r"^\s*port trunk allow-pass vlan\s+(.+)")
REGEX_SHUTDOWN = re.compile(r"^\s*shutdown")
REGEX_ETH_TRUNK_MEMBER = re.compile(r"^\s*eth-trunk\s+(\d+)")
REGEX_IP_ADDRESS = re.compile(r"^\s*ip address\s+([\d\.]+)")

# ==============================================================================
# SECTION 2: PARSING ENGINE
# ==============================================================================

def parse_vrp_config(file_path):
    """
    Parses a Huawei VRP configuration file and returns:
    - meta: hostname, meth_ip, vlanif2_ip
    - interfaces: dict where keys are Interface Names and values are data dicts
    """
    if not os.path.exists(file_path):
        print(f"[Error] Configuration file not found: {file_path}")
        sys.exit(1)

    meta = {
        "hostname": "",
        "meth_ip": "",
        "vlanif2_ip": "",
    }

    interfaces = {}
    current_iface = None

    print(f"[*] Starting lexical analysis of {file_path}...")

    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        for line_num, line in enumerate(f, 1):
            line = line.strip()
            if not line:
                continue

            # Global hostname
            m_sys = REGEX_SYSNAME.search(line)
            if m_sys:
                meta["hostname"] = m_sys.group(1)
                continue

            # Interface block start
            match_start = REGEX_INTERFACE_START.search(line)
            if match_start:
                if current_iface:
                    interfaces[current_iface["name"]] = current_iface

                name = match_start.group(1)
                current_iface = {
                    "name": name,
                    "description": "",
                    "type": "Access",
                    "vlan": "",          # IMPORTANT: default empty (do NOT output vlan 1)
                    "is_shutdown": False,
                    "is_trunk_member": False,
                    "trunk_id": None,
                }
                continue

            # Inside interface block
            if current_iface:
                # block end
                if line == "#":
                    interfaces[current_iface["name"]] = current_iface
                    current_iface = None
                    continue

                # MEth0/0/0 IP and Vlanif2 IP
                m_ip = REGEX_IP_ADDRESS.search(line)
                if m_ip:
                    ip = m_ip.group(1)
                    if current_iface["name"] == "MEth0/0/0":
                        meta["meth_ip"] = ip
                    elif current_iface["name"] == "Vlanif2":
                        meta["vlanif2_ip"] = ip
                    # continue scanning other lines
                    # (no continue)

                # Description
                match_desc = REGEX_DESCRIPTION.search(line)
                if match_desc:
                    current_iface["description"] = match_desc.group(1).strip()
                    continue

                # Link type
                match_type = REGEX_LINK_TYPE.search(line)
                if match_type:
                    current_iface["type"] = match_type.group(1).capitalize()
                    continue

                # Access VLAN (skip vlan 1 -> keep empty)
                match_pvid = REGEX_DEFAULT_VLAN.search(line)
                if match_pvid:
                    vlan_id = match_pvid.group(1)
                    if vlan_id != "1":
                        current_iface["vlan"] = f"Vlan {vlan_id}"
                    else:
                        current_iface["vlan"] = ""
                    continue

                # Trunk VLAN list/range
                match_trunk = REGEX_TRUNK_ALLOW_VLAN.search(line)
                if match_trunk:
                    current_iface["vlan"] = match_trunk.group(1).strip()
                    continue

                # Eth-Trunk membership
                match_member = REGEX_ETH_TRUNK_MEMBER.search(line)
                if match_member:
                    trunk_id = match_member.group(1)
                    current_iface["is_trunk_member"] = True
                    current_iface["trunk_id"] = trunk_id
                    current_iface["type"] = f"Eth-Trunk {trunk_id} Member"
                    current_iface["vlan"] = "See Trunk"
                    continue

                # Shutdown flag
                if REGEX_SHUTDOWN.search(line):
                    current_iface["is_shutdown"] = True

    # Save last interface
    if current_iface:
        interfaces[current_iface["name"]] = current_iface

    print(f"[*] Parsing complete. Identified {len(interfaces)} interface definitions.")
    return meta, interfaces

# ==============================================================================
# SECTION 3: REPORTING ENGINE
# ==============================================================================

def generate_excel_report(meta, interface_data, template_path, output_path):
    """
    Populate Book1.xlsx template:
    - A1 must contain "Hostname:" (already in template)
    - B1 will be: "hostname | meth_ip | vlanif2_ip" (BOLD)
    - For each interface cell found, write:
        row+1: Description (add "Shutdown" if shutdown)
        row+2: Port type
        row+3: Vlan ID (empty if no VLAN)
    """
    if not os.path.exists(template_path):
        print(f"[Error] Template file not found: {template_path}")
        sys.exit(1)

    print(f"[*] Loading Excel template: {template_path}...")

    try:
        wb = load_workbook(template_path)
        ws = wb.active
    except Exception as e:
        print(f"[Error] Failed to load Excel: {e}")
        sys.exit(1)

    # --- Fill B1 with "Hostname | MEth | Vlanif2" (bold) ---
    hostname = meta.get("hostname", "")
    meth_ip = meta.get("meth_ip", "")
    vlanif2_ip = meta.get("vlanif2_ip", "")
    ws["B1"].value = f"{hostname} | {meth_ip} | {vlanif2_ip}"
    ws["B1"].font = Font(bold=True)

    updates_count = 0
    print("[*] Scanning spreadsheet grid for Interface Anchors...")

    for row in ws.iter_rows():
        for cell in row:
            cell_val = str(cell.value).strip() if cell.value else ""
            if cell_val in interface_data:
                target_iface = interface_data[cell_val]

                row_idx = cell.row
                col_idx = cell.column

                # Description row (N+1)
                desc_cell = ws.cell(row=row_idx + 1, column=col_idx)
                desc_text = target_iface["description"].strip() if target_iface["description"] else ""

                # Your rule: if shutdown -> write "Shutdown" in description
                if target_iface["is_shutdown"]:
                    if desc_text:
                        desc_text = f"{desc_text} | Shutdown"
                    else:
                        desc_text = "Shutdown"

                desc_cell.value = desc_text

                # Port type row (N+2) - no "(Down)" here anymore
                type_cell = ws.cell(row=row_idx + 2, column=col_idx)
                type_cell.value = target_iface["type"]

                # VLAN row (N+3) - empty if not configured
                vlan_cell = ws.cell(row=row_idx + 3, column=col_idx)
                vlan_val = target_iface["vlan"]
                if vlan_val == "1":
                    vlan_val = ""
                vlan_cell.value = vlan_val

                updates_count += 1

    print(f"[*] Report generation complete. Populated data for {updates_count} interfaces.")

    try:
        wb.save(output_path)
        print(f"[*] Success! Report saved to: {output_path}")
    except PermissionError:
        print(f"[Error] Permission denied. Is {output_path} open in Excel?")

# ==============================================================================
# SECTION 4: MAIN
# ==============================================================================

def main():
    INPUT_CONFIG = "switch config.txt"
    INPUT_TEMPLATE = "excel.xlsx"
    OUTPUT_FILE = "Huawei_Interface_Audit.xlsx"

    if not os.path.exists(INPUT_CONFIG) and os.path.exists("vrpcfg.cfg"):
        INPUT_CONFIG = "vrpcfg.cfg"

    if not os.path.exists(INPUT_TEMPLATE) and os.path.exists("Book1.xlsx"):
        INPUT_TEMPLATE = "Book1.xlsx"

    print("--- Huawei VRP Interface Audit Tool ---")
    print(f"Target Config: {INPUT_CONFIG}")
    print(f"Target Template: {INPUT_TEMPLATE}")
    print("-" * 40)

    meta, data = parse_vrp_config(INPUT_CONFIG)
    generate_excel_report(meta, data, INPUT_TEMPLATE, OUTPUT_FILE)

if __name__ == "__main__":
    main()
